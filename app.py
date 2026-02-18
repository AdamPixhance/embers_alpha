"""Embers - A local-first daily habits tracker with persistent storage and analytics.

This application provides a clean, modern interface for tracking daily habits with
score aggregation, historical analytics, and streak tracking. Data is stored locally
in SQLite with configuration in Excel.
"""

from __future__ import annotations

import datetime as dt
import importlib.util
import os
import pkgutil
import shutil
import sqlite3
import sys
import threading
import time
from dataclasses import dataclass
from typing import Optional, List, Dict, Tuple

if not hasattr(pkgutil, "find_loader"):
    def _compat_find_loader(module_name: str):
        spec = importlib.util.find_spec(module_name)
        return spec.loader if spec else None

    pkgutil.find_loader = _compat_find_loader

from nicegui import ui, app
from openpyxl import load_workbook

if getattr(sys, "frozen", False):
    APP_DIR = os.path.dirname(sys.executable)
    BUNDLE_DIR = getattr(sys, "_MEIPASS", APP_DIR)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))
    BUNDLE_DIR = APP_DIR

STATIC_DIR = BUNDLE_DIR
DATA_DIR = os.path.join(APP_DIR, "data")
BACKUP_DIR = os.path.join(APP_DIR, "backups")
DB_PATH = os.path.join(DATA_DIR, "habits.db")
XLSX_PATH = os.path.join(DATA_DIR, "Habits.xlsx")
XLSX_TEMPLATE_PATH = os.path.join(BUNDLE_DIR, "data", "Habits.xlsx")

# Serve local files (styles.css, etc.)
app.add_static_files("/static", STATIC_DIR)

CATEGORY_DEFAULTS = {
    "must": 1,
    "good": 2,
    "great": 3,
    "bad": -1,
    "killer": -2,
    "must_avoid": -3,
}

POSITIVE_CATEGORIES = {"must", "good", "great"}
NEGATIVE_CATEGORIES = {"bad", "killer", "must_avoid"}

BADGE_STYLES = {
    "none": {"label": "No Badge", "symbol": "", "color": "rgba(0,0,0,0.08)", "text": "Open"},
    "neg_1": {"label": "Bad", "symbol": "⚠️", "color": "rgba(245, 158, 11, 0.75)", "text": "Bad Day"},
    "neg_2": {"label": "Worse", "symbol": "⛔", "color": "rgba(239, 68, 68, 0.82)", "text": "Worse Day"},
    "neg_3": {"label": "Really Bad", "symbol": "🚨", "color": "rgba(185, 28, 28, 0.9)", "text": "Really Bad Day"},
    "pos_1": {"label": "Bronze", "symbol": "🥉", "color": "rgba(180, 83, 9, 0.82)", "text": "Bronze Day"},
    "pos_2": {"label": "Silver", "symbol": "🥈", "color": "rgba(107, 114, 128, 0.82)", "text": "Silver Day"},
    "pos_3": {"label": "Gold", "symbol": "🥇", "color": "rgba(234, 179, 8, 0.86)", "text": "Gold Day"},
    "pos_4": {"label": "Diamond", "symbol": "💎", "color": "rgba(37, 99, 235, 0.88)", "text": "Diamond Day"},
}

ALLOWED_CATEGORIES = set(CATEGORY_DEFAULTS.keys())
ALLOWED_TYPES = {"check", "number"}
ALLOWED_SCHEDULE_KEYWORDS = {"", "daily", "everyday", "weekdays", "weekday", "weekends", "weekend"}
ALLOWED_ACTIVE_VALUES = {0, 1}

WEEKDAY_MAP = {
    "mon": 0, "monday": 0,
    "tue": 1, "tues": 1, "tuesday": 1,
    "wed": 2, "wednesday": 2,
    "thu": 3, "thur": 3, "thurs": 3, "thursday": 3,
    "fri": 4, "friday": 4,
    "sat": 5, "saturday": 5,
    "sun": 6, "sunday": 6,
}


@dataclass
class Habit:
    habit_id: str
    name: str
    category: str
    weight_override: Optional[float]
    type: str  # check|number
    target: Optional[float]
    schedule: str
    active: int
    notes: str
    label: str = ""
    active_from: Optional[str] = None
    inactive_from: Optional[str] = None

    @property
    def weight(self) -> float:
        if self.weight_override is not None:
            return float(self.weight_override)
        return float(CATEGORY_DEFAULTS.get(self.category.strip().lower(), 0))


def today_local_date() -> str:
    return dt.date.today().isoformat()


def ensure_runtime_files() -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(BACKUP_DIR, exist_ok=True)

    if os.path.exists(XLSX_PATH):
        return

    try:
        if os.path.exists(XLSX_TEMPLATE_PATH):
            shutil.copy2(XLSX_TEMPLATE_PATH, XLSX_PATH)
    except Exception as e:
        print(f"Warning: could not initialize Habits.xlsx from template: {e}")


def parse_excel_iso_date(value: object, row_num: int, field_name: str) -> Optional[str]:
    """Parse optional Excel date/date-string into ISO format (YYYY-MM-DD)."""
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return None
    try:
        return dt.date.fromisoformat(text).isoformat()
    except ValueError as exc:
        raise ValueError(
            f"Row {row_num}: {field_name} must be a valid date in YYYY-MM-DD format"
        ) from exc


def parse_schedule(schedule: str, weekday: int) -> bool:
    """Check if a habit should be scheduled for the given weekday.
    
    Args:
        schedule: Schedule string (daily, weekdays, weekends, or comma-separated day names)
        weekday: Weekday number (0=Monday, 6=Sunday)
    
    Returns:
        True if habit should be shown for this weekday, False otherwise
    """
    s = (schedule or "").strip().lower()
    if s in ("", "daily", "everyday"):
        return True
    if s in ("weekdays", "weekday"):
        return weekday <= 4
    if s in ("weekends", "weekend"):
        return weekday >= 5
    parts = [p.strip() for p in s.split(",") if p.strip()]
    if parts:
        return any(p in WEEKDAY_MAP and WEEKDAY_MAP[p] == weekday for p in parts)
    return True


def validate_schedule_string(raw_schedule: str) -> str:
    """Validate and normalize the schedule string from Excel.

    Accepts: daily-like keywords or comma-separated weekdays (mon,tue,wed...).

    Args:
        raw_schedule: Schedule value from Excel

    Returns:
        Normalized schedule string (lowercase)

    Raises:
        ValueError: If the schedule contains unknown tokens
    """
    schedule = (raw_schedule or "").strip().lower()
    if schedule in ALLOWED_SCHEDULE_KEYWORDS:
        return schedule or "daily"

    parts = [p.strip() for p in schedule.split(",") if p.strip()]
    if not parts:
        return "daily"

    unknown = [p for p in parts if p not in WEEKDAY_MAP]
    if unknown:
        raise ValueError(f"Invalid schedule tokens: {', '.join(unknown)}")

    return ",".join(parts)


def connect() -> sqlite3.Connection:
    """Create and configure a database connection.
    
    Returns:
        SQLite connection with WAL mode and foreign key constraints enabled
    """
    os.makedirs(DATA_DIR, exist_ok=True)
    con = sqlite3.connect(DB_PATH)
    con.execute("PRAGMA journal_mode=WAL;")
    con.execute("PRAGMA foreign_keys=ON;")
    return con


def ensure_column(con: sqlite3.Connection, table: str, column: str, coltype: str) -> None:
    """Add a column to a table if it doesn't already exist (migration helper).
    
    Args:
        con: Database connection
        table: Table name
        column: Column name to add
        coltype: SQL type for the column
    """
    cols = [r[1]
            for r in con.execute(f"PRAGMA table_info({table});").fetchall()]
    if column not in cols:
        con.execute(f"ALTER TABLE {table} ADD COLUMN {column} {coltype};")


def init_db() -> None:
    """Create tables if missing + migrate schema for older DB files."""
    with connect() as con:
        con.execute("""
        CREATE TABLE IF NOT EXISTS habits (
            habit_id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            category TEXT NOT NULL,
            weight_override REAL,
            type TEXT NOT NULL,
            target REAL,
            schedule TEXT NOT NULL,
            active INTEGER NOT NULL DEFAULT 1,
            notes TEXT NOT NULL DEFAULT "",
            updated_at TEXT NOT NULL
        );
        """)

        con.execute("""
        CREATE TABLE IF NOT EXISTS checkins (
            date TEXT NOT NULL,
            habit_id TEXT NOT NULL,
            value REAL NOT NULL,
            checked_at TEXT NOT NULL,
            PRIMARY KEY(date, habit_id),
            FOREIGN KEY(habit_id) REFERENCES habits(habit_id) ON UPDATE CASCADE
        );
        """)

        con.execute("""
        CREATE TABLE IF NOT EXISTS daily_summary (
            date TEXT PRIMARY KEY,
            score REAL NOT NULL,
            max_possible_score REAL NOT NULL,
            pos_score REAL NOT NULL DEFAULT 0,
            neg_score REAL NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        );
        """)

        con.execute("""
        CREATE TABLE IF NOT EXISTS habit_history (
            date TEXT NOT NULL,
            habit_id TEXT NOT NULL,
            completed INTEGER NOT NULL,
            value REAL,
            PRIMARY KEY(date, habit_id),
            FOREIGN KEY(habit_id) REFERENCES habits(habit_id) ON UPDATE CASCADE
        );
        """)

        con.execute("""
        CREATE TABLE IF NOT EXISTS day_status (
            date TEXT PRIMARY KEY,
            completed INTEGER NOT NULL DEFAULT 0,
            completed_at TEXT,
            unlocked_at TEXT,
            updated_at TEXT NOT NULL
        );
        """)

        # migrations for older DBs
        ensure_column(con, "habits", "notes", 'TEXT NOT NULL DEFAULT ""')
        ensure_column(con, "habits", "label", 'TEXT NOT NULL DEFAULT ""')
        ensure_column(con, "habits", "weight_override", "REAL")
        ensure_column(con, "habits", "target", "REAL")
        ensure_column(con, "habits", "active_from", "TEXT")
        ensure_column(con, "habits", "inactive_from", "TEXT")
        ensure_column(con, "daily_summary", "max_negative_score", "REAL NOT NULL DEFAULT 0")

        today = today_local_date()

        # Backfill activation windows for legacy records.
        con.execute(
            """
            UPDATE habits
            SET active_from = COALESCE(
                (SELECT MIN(c.date) FROM checkins c WHERE c.habit_id = habits.habit_id),
                ?
            )
            WHERE active_from IS NULL
            """,
            (today,),
        )
        con.execute(
            """
            UPDATE habits
            SET inactive_from = ?
            WHERE active = 0 AND inactive_from IS NULL
            """,
            (today,),
        )

        con.execute(
            """
            UPDATE daily_summary
            SET max_negative_score = 0
            WHERE max_negative_score IS NULL
            """
        )

        con.commit()


def classify_day_badge(score: float, max_positive_score: float, max_negative_score: float, neg_score: float) -> str:
    if score >= 0:
        ratio = 0.0 if max_positive_score <= 0 else score / max_positive_score
        if ratio >= 0.75:
            return "pos_4"
        if ratio >= 0.50:
            return "pos_3"
        if ratio >= 0.30:
            return "pos_2"
        if ratio >= 0.15:
            return "pos_1"
        return "none"

    neg_ratio = 0.0 if max_negative_score <= 0 else abs(neg_score) / max_negative_score
    if neg_ratio >= 0.50:
        return "neg_3"
    if neg_ratio >= 0.30:
        return "neg_2"
    if neg_ratio >= 0.10:
        return "neg_1"
    return "none"


def get_monday(d: dt.date) -> dt.date:
    return d - dt.timedelta(days=d.weekday())


def get_day_badges_for_range(start_date: str, end_date: str) -> Dict[str, Dict[str, object]]:
    with connect() as con:
        rows = con.execute(
            """
            SELECT ds.date, ds.score, ds.max_possible_score, ds.neg_score, ds.max_negative_score
            FROM daily_summary ds
            JOIN day_status st ON st.date = ds.date
            WHERE st.completed = 1
              AND ds.date >= ?
              AND ds.date <= ?
            ORDER BY ds.date ASC
            """,
            (start_date, end_date),
        ).fetchall()

    data: Dict[str, Dict[str, object]] = {}
    for date_iso, score, max_positive, neg_score, max_negative in rows:
        badge_key = classify_day_badge(
            float(score or 0),
            float(max_positive or 0),
            float(max_negative or 0),
            float(neg_score or 0),
        )
        style = BADGE_STYLES[badge_key]
        data[date_iso] = {
            "badge_key": badge_key,
            "badge_label": style["label"],
            "badge_symbol": style["symbol"],
            "badge_text": style["text"],
            "badge_color": style["color"],
            "score": float(score or 0),
            "max_positive": float(max_positive or 0),
            "max_negative": float(max_negative or 0),
            "neg_score": float(neg_score or 0),
        }
    return data


def get_yearly_badge_counts(year: int) -> Dict[str, int]:
    start = dt.date(year, 1, 1).isoformat()
    end = dt.date(year, 12, 31).isoformat()
    badges = get_day_badges_for_range(start, end)

    counts = {
        "bronze": 0,
        "silver": 0,
        "gold": 0,
        "diamond": 0,
        "bad": 0,
        "worse": 0,
        "really_bad": 0,
    }

    for info in badges.values():
        key = info["badge_key"]
        if key == "pos_1":
            counts["bronze"] += 1
        elif key == "pos_2":
            counts["silver"] += 1
        elif key == "pos_3":
            counts["gold"] += 1
        elif key == "pos_4":
            counts["diamond"] += 1
        elif key == "neg_1":
            counts["bad"] += 1
        elif key == "neg_2":
            counts["worse"] += 1
        elif key == "neg_3":
            counts["really_bad"] += 1

    return counts


def get_tracking_streak() -> int:
    today = dt.date.today()

    with connect() as con:
        completed_dates = {
            row[0]
            for row in con.execute(
                "SELECT date FROM day_status WHERE completed = 1"
            ).fetchall()
        }

    if not completed_dates:
        return 0

    current = today if today.isoformat() in completed_dates else (today - dt.timedelta(days=1))
    streak = 0
    while current.isoformat() in completed_dates:
        streak += 1
        current -= dt.timedelta(days=1)

    return streak


def get_habit_streak_for_day(habit: Habit, as_of_date: str) -> int:
    end_date = dt.date.fromisoformat(as_of_date)
    start_date = end_date - dt.timedelta(days=365)

    with connect() as con:
        checkins_rows = con.execute(
            """
            SELECT date, value
            FROM checkins
            WHERE habit_id = ?
              AND date >= ?
              AND date <= ?
            """,
            (habit.habit_id, start_date.isoformat(), end_date.isoformat()),
        ).fetchall()

    checkins_map = {date_iso: float(val) for date_iso, val in checkins_rows}
    streak = 0
    current = end_date

    while True:
        if habit.active_from and current.isoformat() < habit.active_from:
            break
        if habit.inactive_from and current.isoformat() >= habit.inactive_from:
            current -= dt.timedelta(days=1)
            continue

        if not parse_schedule(habit.schedule, current.weekday()):
            current -= dt.timedelta(days=1)
            continue

        value = checkins_map.get(current.isoformat())
        if not is_habit_completed_for_day(habit, value):
            break

        streak += 1
        current -= dt.timedelta(days=1)

    return streak


def read_habits_from_excel(xlsx_path: str) -> List[Habit]:
    """Load habit definitions from Excel file.
    
    Expects a sheet named 'habits' with columns: habit_id, name, category, type,
    target, schedule, active, and optional: weight_override, label, notes, active_from, inactive_from
    
    Args:
        xlsx_path: Path to Excel file
    
    Returns:
        List of Habit objects parsed from the spreadsheet
    
    Raises:
        ValueError: If 'habits' sheet is missing or required columns are absent
    """
    wb = load_workbook(xlsx_path)
    if "habits" not in wb.sheetnames:
        raise ValueError('Habits.xlsx needs a sheet named "habits".')
    ws = wb["habits"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {name: i for i, name in enumerate(header)}

    required = ["habit_id", "name", "category", "type", "schedule", "active"]
    for r in required:
        if r not in idx:
            raise ValueError(f'Missing column "{r}" in Habits.xlsx.')

    habits: List[Habit] = []
    for row_num, r in enumerate(rows[1:], start=2):
        if r is None or all(v is None or not str(v).strip() for v in r):
            continue

        habit_id = str(r[idx["habit_id"]] or "").strip()
        if not habit_id:
            continue

        name = str(r[idx["name"]] or habit_id).strip()
        category = str(r[idx["category"]] or "").strip().lower()
        if not category:
            raise ValueError(f"Row {row_num}: category is required")
        if category not in ALLOWED_CATEGORIES:
            raise ValueError(f"Row {row_num}: invalid category '{category}'")

        wraw = r[idx.get("weight_override")
                 ] if "weight_override" in idx else None
        weight_override = None
        if wraw is not None and str(wraw).strip() != "":
            try:
                weight_override = float(wraw)
            except ValueError as exc:
                raise ValueError(f"Row {row_num}: weight_override must be numeric") from exc

        type_ = str(r[idx["type"]] or "").strip().lower()
        if not type_:
            raise ValueError(f"Row {row_num}: type is required")
        if type_ not in ALLOWED_TYPES:
            raise ValueError(f"Row {row_num}: invalid type '{type_}'")

        traw = r[idx.get("target")] if "target" in idx else None
        target = None
        if traw is not None and str(traw).strip() != "":
            try:
                target = float(traw)
            except ValueError as exc:
                raise ValueError(f"Row {row_num}: target must be numeric") from exc

        if type_ == "number":
            if target is None or target <= 0:
                raise ValueError(f"Row {row_num}: number type requires target > 0")

        schedule = validate_schedule_string(str(r[idx["schedule"]] or ""))

        active_raw = r[idx["active"]]
        if active_raw is None or str(active_raw).strip() == "":
            raise ValueError(f"Row {row_num}: active is required (0 or 1)")
        try:
            active = int(active_raw)
        except ValueError as exc:
            raise ValueError(f"Row {row_num}: active must be 0 or 1") from exc
        if active not in ALLOWED_ACTIVE_VALUES:
            raise ValueError(f"Row {row_num}: active must be 0 or 1")

        label = ""
        label_col = idx.get("label")
        if label_col is not None:
            label = str(r[label_col] or "").strip()

        notes = str(r[idx.get("notes")] or "").strip(
        ) if "notes" in idx else ""

        active_from = None
        active_from_col = idx.get("active_from")
        if active_from_col is not None:
            active_from = parse_excel_iso_date(r[active_from_col], row_num, "active_from")

        inactive_from = None
        inactive_from_col = idx.get("inactive_from")
        if inactive_from_col is not None:
            inactive_from = parse_excel_iso_date(r[inactive_from_col], row_num, "inactive_from")

        if active_from and inactive_from and inactive_from <= active_from:
            raise ValueError(
                f"Row {row_num}: inactive_from must be after active_from"
            )

        habits.append(Habit(
            habit_id=habit_id,
            name=name,
            category=category,
            weight_override=weight_override,
            type=type_,
            target=target,
            schedule=schedule,
            active=active,
            notes=notes,
            label=label,
            active_from=active_from,
            inactive_from=inactive_from,
        ))

    return habits


def upsert_habits(habits: List[Habit]) -> Tuple[int, int]:
    now = dt.datetime.now().isoformat(timespec="seconds")
    today = today_local_date()
    inserted = 0
    updated = 0

    with connect() as con:
        # safety migration (in case init_db was removed)
        ensure_column(con, "habits", "label", 'TEXT NOT NULL DEFAULT ""')
        ensure_column(con, "habits", "active_from", "TEXT")
        ensure_column(con, "habits", "inactive_from", "TEXT")

        for h in habits:
            cur = con.execute(
                "SELECT active, active_from, inactive_from FROM habits WHERE habit_id = ?",
                (h.habit_id,),
            )
            existing = cur.fetchone()

            if existing is None:
                active_from = h.active_from if h.active == 1 else None
                if h.active == 1 and active_from is None:
                    active_from = today
                inactive_from = h.inactive_from if h.active == 0 else None
                if h.active == 0 and inactive_from is None:
                    inactive_from = today

                if active_from and inactive_from and inactive_from <= active_from:
                    raise ValueError(
                        f"Habit '{h.habit_id}': inactive_from must be after active_from"
                    )
                con.execute(
                    """
                    INSERT INTO habits(
                        habit_id, name, category, weight_override, type, target,
                        schedule, active, notes, label, active_from, inactive_from, updated_at
                    )
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        h.habit_id,
                        h.name,
                        h.category,
                        h.weight_override,
                        h.type,
                        h.target,
                        h.schedule,
                        h.active,
                        h.notes,
                        h.label,
                        active_from,
                        inactive_from,
                        now,
                    ),
                )
                inserted += 1
            else:
                prev_active, prev_active_from, prev_inactive_from = existing
                active_from = prev_active_from
                inactive_from = prev_inactive_from

                if int(prev_active) == 1 and h.active == 0:
                    inactive_from = h.inactive_from or inactive_from or today
                elif int(prev_active) == 0 and h.active == 1:
                    active_from = h.active_from or today
                    inactive_from = h.inactive_from
                elif h.active == 1:
                    if h.active_from is not None:
                        active_from = h.active_from
                    if h.inactive_from is not None:
                        inactive_from = h.inactive_from
                    if active_from is None and (inactive_from is None or inactive_from > today):
                        active_from = today
                    if inactive_from is not None and inactive_from <= today:
                        h.active = 0
                else:
                    if h.inactive_from is not None:
                        inactive_from = h.inactive_from
                    if inactive_from is None:
                        inactive_from = today

                if active_from and inactive_from and inactive_from <= active_from:
                    raise ValueError(
                        f"Habit '{h.habit_id}': inactive_from must be after active_from"
                    )

                con.execute(
                    """
                    UPDATE habits
                    SET
                        name = ?,
                        category = ?,
                        weight_override = ?,
                        type = ?,
                        target = ?,
                        schedule = ?,
                        active = ?,
                        notes = ?,
                        label = ?,
                        active_from = ?,
                        inactive_from = ?,
                        updated_at = ?
                    WHERE habit_id = ?
                    """,
                    (
                        h.name,
                        h.category,
                        h.weight_override,
                        h.type,
                        h.target,
                        h.schedule,
                        h.active,
                        h.notes,
                        h.label,
                        active_from,
                        inactive_from,
                        now,
                        h.habit_id,
                    ),
                )
                updated += 1

        con.commit()

    return inserted, updated


def sync_habits_from_excel(xlsx_path: str) -> Tuple[int, int, int]:
    """Reload habits from Excel and sync with database.
    
    This function:
    - Reads the current Excel file
    - Updates/inserts habits that exist in Excel
    - Marks habits removed from Excel as inactive (preserves history)
    - Validates using the same strict rules as startup
    
    Args:
        xlsx_path: Path to Excel file
    
    Returns:
        Tuple of (inserted, updated, deactivated) counts
    
    Raises:
        ValueError: If Excel validation fails
    """
    if not os.path.exists(xlsx_path):
        return 0, 0, 0
    
    # Read habits from Excel
    habits = read_habits_from_excel(xlsx_path)
    habit_ids_in_excel = {h.habit_id for h in habits}
    
    inserted, updated = upsert_habits(habits)
    
    # Deactivate habits that are no longer in Excel but exist in database
    deactivated = 0
    with connect() as con:
        # Find all active habits in DB that are NOT in Excel
        cur = con.execute(
            """
            SELECT habit_id FROM habits
            WHERE active = 1 AND habit_id NOT IN ({})
            """.format(",".join("?" * len(habit_ids_in_excel)) if habit_ids_in_excel else "NULL"),
            list(habit_ids_in_excel) if habit_ids_in_excel else []
        )
        to_deactivate = [row[0] for row in cur.fetchall()]
        
        if to_deactivate:
            now = dt.datetime.now().isoformat(timespec="seconds")
            today = today_local_date()
            for hid in to_deactivate:
                con.execute(
                    """
                    UPDATE habits
                    SET active = 0,
                        inactive_from = COALESCE(inactive_from, ?),
                        updated_at = ?
                    WHERE habit_id = ?
                    """,
                    (today, now, hid)
                )
                deactivated += 1
        
        con.commit()
    
    return inserted, updated, deactivated


def get_habits_for_day(date_iso: str) -> List[Habit]:
    d = dt.date.fromisoformat(date_iso)
    weekday = d.weekday()

    with connect() as con:
        rows = con.execute("""
                        SELECT habit_id, name, category, weight_override, type, target, schedule, active, notes, label, active_from, inactive_from
            FROM habits
            WHERE (active_from IS NULL OR active_from <= ?)
              AND (inactive_from IS NULL OR ? < inactive_from)
            ORDER BY category ASC, name ASC
        """, (date_iso, date_iso)).fetchall()

    habits = [Habit(*row) for row in rows]
    return [h for h in habits if parse_schedule(h.schedule, weekday)]


def get_checkins_for_day(date_iso: str) -> Dict[str, float]:
    with connect() as con:
        rows = con.execute(
            "SELECT habit_id, value FROM checkins WHERE date = ?",
            (date_iso,),
        ).fetchall()
    return {hid: float(val) for hid, val in rows}


def set_checkin(date_iso: str, habit_id: str, value: float) -> None:
    now = dt.datetime.now().isoformat(timespec="seconds")
    with connect() as con:
        con.execute("""
            INSERT INTO checkins(date, habit_id, value, checked_at)
            VALUES(?,?,?,?)
            ON CONFLICT(date, habit_id) DO UPDATE SET value=excluded.value, checked_at=excluded.checked_at
        """, (date_iso, habit_id, value, now))
        con.commit()


def delete_checkin(date_iso: str, habit_id: str) -> None:
    with connect() as con:
        con.execute(
            "DELETE FROM checkins WHERE date = ? AND habit_id = ?",
            (date_iso, habit_id),
        )
        con.commit()


def compute_score(habits: List[Habit], checkins: Dict[str, float]) -> Tuple[float, float]:
    """Calculate the daily score based on completed habits and their weights.
    
    For each habit:
    - If type='check': add weight if checked (val != 0)
    - If type='number' with target: add weight * (val/target) clamped to [0, 1]
    - Positive weights contribute to max_score; negative weights do not
    
    Args:
        habits: List of habits to evaluate
        checkins: Dict mapping habit_id -> value (checked/unchecked)
    
    Returns:
        Tuple of (daily_score, max_possible_score)
    """
    score = 0.0
    max_score = 0.0

    for h in habits:
        w = h.weight
        val = checkins.get(h.habit_id)

        if h.type == "check":
            if val is not None and val != 0:
                score += w
        elif h.target and h.target > 0:
            score += w * min(max((val or 0.0) / h.target, 0.0), 1.0)
        elif (val or 0.0) > 0:
            score += w

        if w > 0:
            max_score += w

    return score, max_score


def display_label(h: Habit) -> str:
    return h.label.strip() if (h.label and h.label.strip()) else h.name.strip()


def is_habit_completed_for_day(habit: Habit, value: Optional[float]) -> bool:
    if habit.type == "check":
        return value is not None and value != 0
    if habit.target and habit.target > 0:
        return (value or 0.0) >= habit.target
    return (value or 0.0) > 0


def archive_daily_data(date_iso: str, habits: List[Habit], checkins: Dict[str, float]) -> None:
    """Archive today's score and checkins for historical tracking."""
    score, max_score = compute_score(habits, checkins)
    now = dt.datetime.now().isoformat(timespec="seconds")
    
    # Separate positive and negative scores
    pos_score = 0.0
    neg_score = 0.0
    max_negative_score = sum(abs(h.weight) for h in habits if h.weight < 0)
    for h in habits:
        w = h.weight
        val = checkins.get(h.habit_id)
        if h.type == "check" and val is not None and val != 0:
            if w > 0:
                pos_score += w
            else:
                neg_score += w
    
    with connect() as con:
        con.execute("""
            INSERT INTO daily_summary(date, score, max_possible_score, pos_score, neg_score, max_negative_score, created_at)
            VALUES(?,?,?,?,?,?,?)
            ON CONFLICT(date) DO UPDATE SET 
                score=excluded.score,
                max_possible_score=excluded.max_possible_score,
                pos_score=excluded.pos_score,
                neg_score=excluded.neg_score,
                max_negative_score=excluded.max_negative_score
        """, (date_iso, score, max_score, pos_score, neg_score, max_negative_score, now))
        con.commit()


def save_habit_history_snapshot(date_iso: str, habits: List[Habit], checkins: Dict[str, float]) -> None:
    with connect() as con:
        con.execute("DELETE FROM habit_history WHERE date = ?", (date_iso,))
        for h in habits:
            val = checkins.get(h.habit_id)
            completed = 1 if is_habit_completed_for_day(h, val) else 0
            con.execute(
                """
                INSERT INTO habit_history(date, habit_id, completed, value)
                VALUES(?,?,?,?)
                """,
                (date_iso, h.habit_id, completed, val),
            )
        con.commit()


def is_day_completed(date_iso: str) -> bool:
    with connect() as con:
        row = con.execute(
            "SELECT completed FROM day_status WHERE date = ?",
            (date_iso,),
        ).fetchone()
    return bool(row and int(row[0]) == 1)


def set_day_completed(date_iso: str, completed: bool) -> None:
    now = dt.datetime.now().isoformat(timespec="seconds")
    completed_at = now if completed else None
    unlocked_at = now if not completed else None
    with connect() as con:
        con.execute(
            """
            INSERT INTO day_status(date, completed, completed_at, unlocked_at, updated_at)
            VALUES(?,?,?,?,?)
            ON CONFLICT(date) DO UPDATE SET
                completed=excluded.completed,
                completed_at=excluded.completed_at,
                unlocked_at=excluded.unlocked_at,
                updated_at=excluded.updated_at
            """,
            (date_iso, 1 if completed else 0, completed_at, unlocked_at, now),
        )
        con.commit()


def get_latest_open_day(before_date: Optional[str] = None, exclude_date: Optional[str] = None) -> Optional[str]:
    with connect() as con:
        rows = con.execute(
            """
            SELECT c.date
            FROM checkins c
            LEFT JOIN day_status ds ON ds.date = c.date
            WHERE c.value != 0
              AND COALESCE(ds.completed, 0) = 0
            GROUP BY c.date
            ORDER BY c.date DESC
            """
        ).fetchall()

    for (date_iso,) in rows:
        if before_date is not None and date_iso >= before_date:
            continue
        if exclude_date is not None and date_iso == exclude_date:
            continue
        return date_iso
    return None


def get_day_snapshot(date_iso: str) -> Dict[str, object]:
    habits = get_habits_for_day(date_iso)
    checkins = get_checkins_for_day(date_iso)
    score, max_score = compute_score(habits, checkins)

    done: List[str] = []
    missed: List[str] = []
    for h in habits:
        val = checkins.get(h.habit_id)
        if is_habit_completed_for_day(h, val):
            done.append(display_label(h))
        else:
            missed.append(display_label(h))

    return {
        "date": date_iso,
        "score": score,
        "max_score": max_score,
        "done": sorted(done),
        "missed": sorted(missed),
        "habits": habits,
        "checkins": checkins,
    }


def complete_day(date_iso: str) -> Dict[str, object]:
    snapshot = get_day_snapshot(date_iso)
    habits = snapshot["habits"]
    checkins = snapshot["checkins"]
    archive_daily_data(date_iso, habits, checkins)
    save_habit_history_snapshot(date_iso, habits, checkins)
    set_day_completed(date_iso, True)
    return snapshot


def unlock_day(date_iso: str) -> None:
    set_day_completed(date_iso, False)


def get_streak() -> Tuple[int, bool]:
    """Get current streak and whether it's positive.
    Returns (streak_length, is_positive)
    """
    with connect() as con:
        rows = con.execute("""
            SELECT ds.date, ds.score
            FROM daily_summary ds
            WHERE EXISTS (
                SELECT 1 FROM checkins c
                WHERE c.date = ds.date AND c.value != 0
            )
            ORDER BY ds.date DESC
            LIMIT 365
        """).fetchall()
    
    if not rows:
        return 0, True
    
    score_by_date = {date_str: score for date_str, score in rows}
    today = dt.date.today()
    today_str = today.isoformat()
    if today_str not in score_by_date:
        today_checkins = get_checkins_for_day(today_str)
        if any(val != 0 for val in today_checkins.values()):
            today_habits = get_habits_for_day(today_str)
            today_score, _ = compute_score(today_habits, today_checkins)
            score_by_date[today_str] = today_score
        else:
            return 0, True

    is_positive = score_by_date[today_str] >= 0

    # Count consecutive days with same sign; empty days break the streak
    streak = 0
    current = today
    while True:
        date_str = current.isoformat()
        score = score_by_date.get(date_str)
        if score is None:
            break
        if (score >= 0) != is_positive:
            break
        streak += 1
        current -= dt.timedelta(days=1)

    return streak, is_positive


def get_habit_stats() -> Dict[str, Dict]:
    """Get completion stats for all habits.
    Returns dict with habit_id -> {name, completed_count, total_days, completion_rate}
    """
    with connect() as con:
        # Get all habits
        habit_rows = con.execute("""
            SELECT habit_id, name, category, type, target, schedule, active_from, inactive_from
            FROM habits
            WHERE active = 1
        """).fetchall()
        
        stats = {}
        for habit_id, name, category, type_, target, schedule, active_from, inactive_from in habit_rows:
            completed = con.execute(
                "SELECT COUNT(*) FROM checkins WHERE habit_id = ? AND value != 0",
                (habit_id,)
            ).fetchone()[0]
            
            total = con.execute(
                "SELECT COUNT(DISTINCT date) FROM checkins WHERE habit_id = ?",
                (habit_id,)
            ).fetchone()[0]
            
            rate = (completed / total * 100) if total > 0 else 0
            stats[habit_id] = {
                "habit_id": habit_id,
                "name": name,
                "category": (category or "").strip().lower(),
                "type": type_,
                "target": target,
                "schedule": schedule,
                "active_from": active_from,
                "inactive_from": inactive_from,
                "completed": completed,
                "total_days": total,
                "completion_rate": rate
            }
    
    return stats


def get_date_range_stats(start_date: str, end_date: str) -> Dict:
    """Get stats for a date range."""
    with connect() as con:
        rows = con.execute("""
            SELECT ds.date, ds.score
            FROM daily_summary ds
            WHERE ds.date >= ? AND ds.date <= ?
              AND EXISTS (
                  SELECT 1 FROM checkins c
                  WHERE c.date = ds.date AND c.value != 0
              )
            ORDER BY ds.date ASC
        """, (start_date, end_date)).fetchall()
    
    if not rows:
        return {"avg_score": 0, "best_day": None, "worst_day": None, "total_days": 0}
    
    scores = [score for _, score in rows]
    best_day = max(rows, key=lambda x: x[1])
    worst_day = min(rows, key=lambda x: x[1])
    
    return {
        "avg_score": sum(scores) / len(scores),
        "best_day": best_day,
        "worst_day": worst_day,
        "total_days": len(rows),
        "days": rows
    }


def backup_excel() -> None:
    """Create a timestamped backup of the Excel configuration file.
    
    Silently fails if backup cannot be created (does not affect app operation).
    """
    if os.path.exists(XLSX_PATH):
        os.makedirs(BACKUP_DIR, exist_ok=True)
        backup_path = os.path.join(
            BACKUP_DIR,
            f"Habits_backup_{dt.date.today().isoformat()}.xlsx",
        )
        try:
            shutil.copy2(XLSX_PATH, backup_path)
        except Exception as e:
            print(f"Backup failed: {e}")


def main() -> None:
    ensure_runtime_files()
    init_db()

    # Load external stylesheet (with cache busting)
    css_path = os.path.join(STATIC_DIR, "styles.css")
    if os.path.exists(css_path):
        v = int(os.path.getmtime(css_path))
        ui.add_head_html(
            f'<link rel="stylesheet" href="/static/styles.css?v={v}">')

    # initial sync
    if os.path.exists(XLSX_PATH):
        habits = read_habits_from_excel(XLSX_PATH)
        upsert_habits(habits)

    # State for current viewing date
    current_date = {"value": today_local_date()}
    
    def get_habits_for_current_date():
        return get_habits_for_day(current_date["value"])
    
    habits_for_today: List[Habit] = get_habits_for_current_date()
    checkins = get_checkins_for_day(current_date["value"])

    def get_habit_streak_map_for_date(date_iso: str) -> Dict[str, int]:
        streak_map: Dict[str, int] = {}
        for habit in get_habits_for_day(date_iso):
            streak_map[habit.habit_id] = get_habit_streak_for_day(habit, date_iso)
        return streak_map

    last_save = {"value": None}
    excel_sync_state = {
        "last_mtime": os.path.getmtime(XLSX_PATH) if os.path.exists(XLSX_PATH) else None,
        "needs_refresh": False
    }
    shutdown_state = {"clients": 0, "timer": None}
    shutdown_grace_seconds = 8
    startup_prompt_state = {"shown": False}

    def check_and_sync_excel() -> bool:
        """Check if Excel file has been modified and sync if needed.
        
        Returns:
            True if habits were updated, False otherwise
        """
        if not os.path.exists(XLSX_PATH):
            return False
        
        try:
            current_mtime = os.path.getmtime(XLSX_PATH)
            if excel_sync_state["last_mtime"] is None:
                excel_sync_state["last_mtime"] = current_mtime
                return False
            
            # File has been modified
            if current_mtime != excel_sync_state["last_mtime"]:
                excel_sync_state["last_mtime"] = current_mtime
                inserted, updated, deactivated = sync_habits_from_excel(XLSX_PATH)
                if inserted > 0 or updated > 0 or deactivated > 0:
                    excel_sync_state["needs_refresh"] = True
                    return True
        except Exception as e:
            print(f"Warning: Excel sync check failed: {e}")
        
        return False

    def save_and_backup() -> None:
        now = dt.datetime.now()
        if last_save["value"] and (now - last_save["value"]).total_seconds() < 2:
            return
        last_save["value"] = now

        if current_date["value"] < today_local_date():
            old_habits = get_habits_for_current_date()
            old_checkins = get_checkins_for_day(current_date["value"])
            archive_daily_data(current_date["value"], old_habits, old_checkins)

        backup_excel()

    def _shutdown_if_idle() -> None:
        if shutdown_state["clients"] > 0:
            return
        save_and_backup()
        sys.exit(0)

    def _cancel_shutdown_timer() -> None:
        timer = shutdown_state.get("timer")
        if timer is not None:
            timer.cancel()
            shutdown_state["timer"] = None

    def _schedule_shutdown_timer() -> None:
        _cancel_shutdown_timer()
        timer = threading.Timer(shutdown_grace_seconds, _shutdown_if_idle)
        timer.daemon = True
        shutdown_state["timer"] = timer
        timer.start()

    goodbye_dialog = ui.dialog()
    with goodbye_dialog:
        with ui.card().classes("goodbye-card"):
            ui.label("Goodbye").classes("goodbye-title")
            ui.label("Saving your progress... You can close this window.").classes("goodbye-sub")

    snapshot_dialog = ui.dialog()
    with snapshot_dialog:
        with ui.card().style("min-width: 560px; max-width: 92vw;"):
            snapshot_title = ui.label("Day Snapshot").style("font-size: 1.1rem; font-weight: 700;")
            snapshot_score = ui.label("")
            with ui.row().classes("w-full").style("gap: 24px; align-items: flex-start;"):
                with ui.column().style("flex: 1;"):
                    ui.label("Done").style("font-weight: 600;")
                    snapshot_done_col = ui.column().classes("w-full")
                with ui.column().style("flex: 1;"):
                    ui.label("Not Done").style("font-weight: 600;")
                    snapshot_missed_col = ui.column().classes("w-full")
            ui.button("Close", on_click=snapshot_dialog.close).props("flat")

    open_day_dialog = ui.dialog()
    with open_day_dialog:
        with ui.card().style("min-width: 520px; max-width: 92vw;"):
            open_day_title = ui.label("Unfinished Day Found").style("font-size: 1.1rem; font-weight: 700;")
            open_day_msg = ui.label("")
            with ui.row().classes("w-full").style("justify-content: flex-end; gap: 8px;"):
                open_day_continue_btn = ui.button("Continue That Day").props("flat")
                open_day_finalize_btn = ui.button("Finalize That Day").props("color=positive")

    open_day_target = {"value": None}

    def show_snapshot_dialog(snapshot: Dict[str, object]) -> None:
        snapshot_title.text = f"Day Snapshot — {snapshot['date']}"
        snapshot_score.text = f"Score: {snapshot['score']:.1f} / {snapshot['max_score']:.1f}"

        snapshot_done_col.clear()
        with snapshot_done_col:
            done_list = snapshot["done"]
            if done_list:
                for name in done_list:
                    ui.label(f"✓ {name}")
            else:
                ui.label("No habits completed.")

        snapshot_missed_col.clear()
        with snapshot_missed_col:
            missed_list = snapshot["missed"]
            if missed_list:
                for name in missed_list:
                    ui.label(f"• {name}")
            else:
                ui.label("Everything completed.")

        snapshot_dialog.open()

    def open_unresolved_day_prompt(date_iso: str) -> None:
        open_day_target["value"] = date_iso
        open_day_msg.text = (
            f"{date_iso} has check-ins and is still open. "
            "Finalize it now or continue editing that day before opening another one."
        )
        open_day_dialog.open()

    def go_to_day(date_iso: str) -> None:
        current_date["value"] = date_iso
        refresh_header()
        render_habits.refresh()

    def handle_continue_open_day() -> None:
        target = open_day_target["value"]
        open_day_dialog.close()
        if target:
            go_to_day(target)

    def handle_finalize_open_day() -> None:
        target = open_day_target["value"]
        open_day_dialog.close()
        if not target:
            return
        snapshot = complete_day(target)
        ui.notify(f"Day {target} completed and locked.", type="positive")
        if current_date["value"] == target:
            refresh_header()
            render_habits.refresh()
        show_snapshot_dialog(snapshot)

    open_day_continue_btn.on_click(handle_continue_open_day)
    open_day_finalize_btn.on_click(handle_finalize_open_day)

    @app.post("/save_on_exit")
    def save_on_exit() -> Dict[str, bool]:
        save_and_backup()
        return {"ok": True}

    # Header with date navigation
    with ui.column().classes("w-full").style("padding:0;"):
        # Date navigation row with flexbox layout
        with ui.row().classes("w-full date-nav-row").style("align-items: center; justify-content: flex-start;"):
            # Left: Previous/Today/Next
            with ui.row().classes("date-nav-group").style("flex: 0 0 auto;"):
                prev_btn = ui.button("← Previous").props("flat")
                today_btn = ui.button("Today").props("flat")
                next_btn = ui.button("Next →").props("flat")
            
            # Center: Date (flex-grow to center it)
            date_label = ui.label(current_date["value"]).classes("date-nav-date").style("flex: 1; text-align: center;")
            
            # Right: Action buttons grouped together
            with ui.row().classes("button-group-right").style("gap: 8px; flex: 0 0 auto; margin-left: auto;"):
                complete_btn = ui.button("Complete Day").props("color=positive")
                unlock_btn = ui.button("Unlock Day").props("color=warning")
                close_btn = ui.button("Close App").props("flat").style("color: var(--text-secondary) !important;")
        
        # Tabbed interface
        with ui.tabs() as tabs:
            tab_daily = ui.tab("Daily Habits")
            tab_stats = ui.tab("Stats")
        
        with ui.tab_panels(tabs, value=tab_daily):
            # Daily Habits Tab
            with ui.tab_panel(tab_daily):
                with ui.column().classes("w-full center-wrap"):
                    with ui.column().classes("app-shell"):
                        # Score header
                        with ui.column().classes("w-full header-container"):
                            ui.label("Daily Habits").classes("header-title")
                            ui.label(current_date["value"]).classes("header-sub")

                            score_label = ui.label().classes("score-display")
                            tracking_streak_label = ui.label().style("font-weight: 600;")
                            yearly_badges_label = ui.label().style("color: var(--text-secondary);")
                            bar_container = ui.element("div").classes("score-bar-positive")
                            with bar_container:
                                bar = ui.linear_progress(value=0.0).props("size=16px rounded").props("instant-feedback")

                            def refresh_header() -> None:
                                nonlocal habits_for_today, checkins
                                
                                # Check if Excel file has changed and sync if needed
                                if check_and_sync_excel():
                                    excel_sync_state["needs_refresh"] = False
                                    # If viewing today, refresh the habit list immediately
                                    if current_date["value"] == today_local_date():
                                        render_habits.refresh()
                                
                                habits_for_today = get_habits_for_current_date()
                                checkins = get_checkins_for_day(current_date["value"])
                                s, m = compute_score(habits_for_today, checkins)
                                
                                # Calculate percentage and determine bar style
                                if s >= 0:
                                    ratio = 0.0 if m <= 0 else max(min(s / m, 1.0), 0.0)
                                    bar_container.classes(remove="score-bar-negative")
                                    bar_container.classes(add="score-bar-positive")
                                    percentage = int(ratio * 100)
                                else:
                                    max_negative = sum(abs(h.weight) for h in habits_for_today if h.weight < 0)
                                    ratio = 0.0 if max_negative <= 0 else max(min(abs(s) / max_negative, 1.0), 0.0)
                                    bar_container.classes(remove="score-bar-positive")
                                    bar_container.classes(add="score-bar-negative")
                                    percentage = -int(ratio * 100)
                                
                                score_label.text = f"Score: {s:.1f} / {m:.1f}  •  Progress: {percentage}%"
                                bar.value = ratio
                                date_label.text = current_date["value"]

                                tracking = get_tracking_streak()
                                year_counts = get_yearly_badge_counts(dt.date.today().year)
                                tracking_streak_label.text = f"Tracking Streak: {tracking} day(s)"
                                yearly_badges_label.text = (
                                    f"Yearly Badges — 🥉 {year_counts['bronze']}  🥈 {year_counts['silver']}  "
                                    f"🥇 {year_counts['gold']}  💎 {year_counts['diamond']}  "
                                    f"⚠️ {year_counts['bad']}  ⛔ {year_counts['worse']}  🚨 {year_counts['really_bad']}"
                                )

                                locked = is_day_completed(current_date["value"])
                                complete_btn.visible = not locked
                                unlock_btn.visible = locked

                        refresh_header()

                    @ui.refreshable
                    def render_habits():
                        nonlocal habits_for_today, checkins
                        habits_for_today = get_habits_for_current_date()
                        checkins = get_checkins_for_day(current_date["value"])
                        streak_map = get_habit_streak_map_for_date(current_date["value"])
                        day_locked = is_day_completed(current_date["value"])

                        if day_locked:
                            ui.label("This day is completed and locked. Unlock to edit.").style("font-weight: 600; margin: 8px 0;")

                        pos_categories = [("must", "Must"), ("good", "Good"), ("great", "Great")]
                        neg_categories = [("bad", "Bad"), ("killer", "Killer"), ("must_avoid", "Must Avoid")]

                        def render_category_group(kind: str, categories: list) -> None:
                            with ui.element("div").classes("pyramid-layers-container"):
                                for cat, label in categories:
                                    items = [h for h in habits_for_today if h.category ==
                                             cat and h.type == "check"]
                                    
                                    with ui.element("div").classes("category-group"):
                                        ui.label(label).classes("category-group-title")
                                        
                                        with ui.card().classes("pyramid-layer"):
                                            if not items:
                                                ui.element("div").classes("layer-pad")
                                            else:
                                                with ui.element("div").classes("habit-row"):
                                                    for h in items:
                                                        done_state = (checkins.get(h.habit_id, 0.0) != 0.0)

                                                        streak_value = streak_map.get(h.habit_id, 0)
                                                        streak_symbol = "🔥" if h.category in POSITIVE_CATEGORIES else "⚠️"
                                                        streak_suffix = f"  {streak_symbol}{streak_value}" if streak_value >= 2 else ""

                                                        btn = ui.button(f"{display_label(h)}{streak_suffix}")
                                                        btn.props('flat no-caps')
                                                        btn.classes("habit-btn")
                                                        btn.classes(add="is-pos" if kind == "pos" else "is-neg")
                                                        btn.classes(add=f"cat-{h.category.strip().lower()}")
                                                        if done_state:
                                                            btn.classes(add="is-done")

                                                        def make_handler(habit_id=h.habit_id):
                                                            def handler():
                                                                if is_day_completed(current_date["value"]):
                                                                    ui.notify("This day is locked. Click Unlock Day to edit.", type="warning")
                                                                    return

                                                                blocking_open_day = get_latest_open_day(exclude_date=current_date["value"])
                                                                if blocking_open_day:
                                                                    ui.notify(
                                                                        f"Finish {blocking_open_day} first. Only one open day is allowed.",
                                                                        type="warning",
                                                                    )
                                                                    open_unresolved_day_prompt(blocking_open_day)
                                                                    return

                                                                current = (get_checkins_for_day(
                                                                    current_date["value"]).get(habit_id, 0.0) != 0.0)
                                                                if current:
                                                                    delete_checkin(current_date["value"], habit_id)
                                                                else:
                                                                    set_checkin(current_date["value"], habit_id, 1.0)
                                                                refresh_header()
                                                                render_habits.refresh()
                                                            return handler

                                                        btn.on("click", make_handler())
                                                        if day_locked:
                                                            btn.disable()

                        with ui.element("section").classes("pyramid-section pos-section"):
                            render_category_group("pos", pos_categories)

                        with ui.element("section").classes("pyramid-section neg-section"):
                            render_category_group("neg", neg_categories)

                    render_habits()
            
            # Stats Tab
            with ui.tab_panel(tab_stats):
                with ui.column().classes("w-full center-wrap"):
                    with ui.column().classes("app-shell").style("padding: 16px;"):
                        ui.label("Statistics").classes("header-title")
                        
                        # Tracking streak + yearly badge counters
                        with ui.row().classes("w-full").style("gap: 16px;"):
                            ui.label(f"Tracking streak: {get_tracking_streak()} days").style("font-size: 1.1rem; font-weight: 600;")

                        yearly = get_yearly_badge_counts(dt.date.today().year)
                        ui.label(
                            f"{dt.date.today().year} badges — 🥉 {yearly['bronze']}  🥈 {yearly['silver']}  🥇 {yearly['gold']}  💎 {yearly['diamond']}  "
                            f"⚠️ {yearly['bad']}  ⛔ {yearly['worse']}  🚨 {yearly['really_bad']}"
                        ).style("margin-bottom: 8px;")
                        
                        # Date range stats
                        today = dt.date.today()
                        last_7 = (today - dt.timedelta(days=6)).isoformat()
                        last_30 = (today - dt.timedelta(days=29)).isoformat()
                        
                        stats_7 = get_date_range_stats(last_7, today.isoformat())
                        stats_30 = get_date_range_stats(last_30, today.isoformat())
                        
                        with ui.row().classes("w-full").style("gap: 16px;"):
                            with ui.card().classes("stat-card").style("flex: 1;"):
                                ui.label("7-Day Average").classes("stat-label")
                                ui.label(f"{stats_7['avg_score']:.1f}").classes("stat-value")
                            
                            with ui.card().classes("stat-card").style("flex: 1;"):
                                ui.label("30-Day Average").classes("stat-label")
                                ui.label(f"{stats_30['avg_score']:.1f}").classes("stat-value")
                        
                        def render_badge_grid(title: str, start: dt.date, end: dt.date, columns: int = 7) -> None:
                            ui.label(title).style("font-size: 0.95rem; font-weight: 600; margin-top: 16px;")
                            badges = get_day_badges_for_range(start.isoformat(), end.isoformat())

                            with ui.element("div").style(
                                f"display: grid; grid-template-columns: repeat({columns}, 1fr); gap: 4px; margin-top: 10px;"
                            ):
                                current = start
                                while current <= end:
                                    date_str = current.isoformat()
                                    info = badges.get(date_str)
                                    if info:
                                        color = info["badge_color"]
                                        symbol = info["badge_symbol"] or "•"
                                        label = info["badge_text"]
                                        score = info["score"]
                                    else:
                                        color = BADGE_STYLES["none"]["color"]
                                        symbol = ""
                                        label = "Open"
                                        score = 0.0

                                    cell = ui.element("div").style(
                                        f"width: 28px; height: 28px; background-color: {color}; border-radius: 4px; "
                                        "display: flex; align-items: center; justify-content: center; font-size: 0.7rem;"
                                    )
                                    with cell:
                                        ui.label(symbol).style("font-size: 0.7rem;")
                                    cell.tooltip(f"{date_str}: {label} • Score {score:.1f}")
                                    current += dt.timedelta(days=1)

                        today_date = dt.date.today()
                        week_start = get_monday(today_date)
                        week_end = week_start + dt.timedelta(days=6)
                        month_start = dt.date(today_date.year, today_date.month, 1)
                        next_month_start = (
                            dt.date(today_date.year + 1, 1, 1)
                            if today_date.month == 12
                            else dt.date(today_date.year, today_date.month + 1, 1)
                        )
                        month_end = next_month_start - dt.timedelta(days=1)
                        year_start = dt.date(today_date.year, 1, 1)
                        year_end = dt.date(today_date.year, 12, 31)

                        render_badge_grid("Weekly Grid", week_start, week_end, columns=7)
                        render_badge_grid("Monthly Grid", month_start, month_end, columns=7)
                        render_badge_grid("Yearly Grid", year_start, year_end, columns=14)
                        
                        # Habit performance table
                        ui.label("Habit Performance").style("font-size: 0.95rem; font-weight: 600; margin-top: 24px;")
                        
                        habit_stats = get_habit_stats()
                        if habit_stats:
                            category_order = [
                                ("must", "Must"),
                                ("good", "Good"),
                                ("great", "Great"),
                                ("bad", "Bad"),
                                ("killer", "Killer"),
                                ("must_avoid", "Must Avoid"),
                            ]
                            pos_categories = {"must", "good", "great"}
                            grouped: Dict[str, List[Dict]] = {}
                            for stat in habit_stats.values():
                                grouped.setdefault(stat["category"], []).append(stat)

                            with ui.column().classes("w-full stats-groups"):
                                for cat_key, label in category_order:
                                    items = grouped.get(cat_key, [])
                                    if not items:
                                        continue

                                    title_class = "stats-group-title is-pos" if cat_key in pos_categories else "stats-group-title is-neg"
                                    ui.label(label).classes(title_class)

                                    for stat in sorted(items, key=lambda s: (-s["completion_rate"], s["name"])):
                                        name = stat["name"]
                                        rate = stat["completion_rate"]

                                        with ui.row().classes("w-full stats-row"):
                                            streak = get_habit_streak_for_day(
                                                Habit(
                                                    habit_id=stat.get("habit_id", ""),
                                                    name=stat["name"],
                                                    category=stat["category"],
                                                    weight_override=None,
                                                    type=stat.get("type", "check"),
                                                    target=stat.get("target"),
                                                    schedule=stat.get("schedule", "daily"),
                                                    active=1,
                                                    notes="",
                                                    active_from=stat.get("active_from"),
                                                    inactive_from=stat.get("inactive_from"),
                                                ),
                                                today_local_date(),
                                            ) if stat.get("habit_id") else 0

                                            symbol = "🔥" if stat["category"] in POSITIVE_CATEGORIES else "⚠️"
                                            streak_suffix = f"  {symbol}{streak}" if streak >= 2 else ""
                                            ui.label(f"{name}{streak_suffix}").classes("stats-name")
                                            with ui.linear_progress(value=rate / 100.0).props("instant-feedback").classes("stats-bar"):
                                                pass
                                            ui.label(f"{rate:.0f}%").classes("stats-rate")

    # Date navigation handlers
    def go_previous_day():
        current_date_obj = dt.date.fromisoformat(current_date["value"])
        new_date = (current_date_obj - dt.timedelta(days=1)).isoformat()
        go_to_day(new_date)
    
    def go_next_day():
        current_date_obj = dt.date.fromisoformat(current_date["value"])
        new_date = (current_date_obj + dt.timedelta(days=1)).isoformat()
        # Only allow going to today or earlier
        if new_date <= today_local_date():
            # Archive current day if moving forward
            if current_date["value"] < today_local_date():
                old_habits = get_habits_for_current_date()
                old_checkins = get_checkins_for_day(current_date["value"])
                archive_daily_data(current_date["value"], old_habits, old_checkins)
            go_to_day(new_date)
    
    def go_today():
        if current_date["value"] < today_local_date():
            # Archive the day we're leaving
            old_habits = get_habits_for_current_date()
            old_checkins = get_checkins_for_day(current_date["value"])
            archive_daily_data(current_date["value"], old_habits, old_checkins)
        go_to_day(today_local_date())

    def on_complete_day() -> None:
        date_iso = current_date["value"]
        if is_day_completed(date_iso):
            ui.notify("This day is already completed.", type="warning")
            refresh_header()
            render_habits.refresh()
            return

        blocking_open_day = get_latest_open_day(exclude_date=date_iso)
        if blocking_open_day:
            ui.notify(
                f"Finish {blocking_open_day} first. Only one open day is allowed.",
                type="warning",
            )
            open_unresolved_day_prompt(blocking_open_day)
            return

        snapshot = complete_day(date_iso)
        ui.notify(f"Day {date_iso} completed and locked.", type="positive")
        refresh_header()
        render_habits.refresh()
        show_snapshot_dialog(snapshot)

    def on_unlock_day() -> None:
        date_iso = current_date["value"]
        unlock_day(date_iso)
        ui.notify(f"Day {date_iso} unlocked. You can edit it again.", type="warning")
        refresh_header()
        render_habits.refresh()
    
    prev_btn.on_click(go_previous_day)
    next_btn.on_click(go_next_day)
    today_btn.on_click(go_today)
    complete_btn.on_click(on_complete_day)
    unlock_btn.on_click(on_unlock_day)
    
    def on_close():
        save_and_backup()
        goodbye_dialog.open()
        ui.notify("App data saved. Closing...", type="positive")
        ui.run_javascript(
            "setTimeout(() => { try { window.close(); } catch (e) {} }, 800);"
        )
        ui.timer(1.2, lambda: sys.exit(0), once=True)
    
    close_btn.on_click(on_close)

    def register_exit_hook() -> None:
        ui.run_javascript(
            """
            if (!window.__embersExitHook) {
                window.__embersExitHook = true;
                window.addEventListener('beforeunload', () => {
                    try { navigator.sendBeacon('/save_on_exit'); } catch (e) {}
                });
            }
            """
        )

    def prompt_stale_open_day_once() -> None:
        if startup_prompt_state["shown"]:
            return
        startup_prompt_state["shown"] = True
        stale = get_latest_open_day(before_date=today_local_date())
        if stale:
            open_unresolved_day_prompt(stale)

    app.on_connect(register_exit_hook)
    app.on_connect(prompt_stale_open_day_once)
    app.on_connect(lambda: (_cancel_shutdown_timer(), shutdown_state.__setitem__("clients", shutdown_state["clients"] + 1)))
    app.on_disconnect(lambda: (shutdown_state.__setitem__("clients", max(shutdown_state["clients"] - 1, 0)), _schedule_shutdown_timer()))
    
    # Background Excel sync check (every 2 seconds)
    def _background_sync_check() -> None:
        try:
            if check_and_sync_excel() and excel_sync_state["needs_refresh"]:
                excel_sync_state["needs_refresh"] = False
                # Refresh habits if viewing today or currently showing daily habit  
                render_habits.refresh()
        except Exception as e:
            print(f"Background sync error: {e}")
    
    ui.timer(2.0, _background_sync_check)
    
    # Graceful shutdown
    ui.run(title="Embers — Daily Habits", port=8080, reload=False, show=False)


if __name__ == "__main__":
    main()
